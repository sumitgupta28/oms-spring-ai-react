#!/usr/bin/env python3
"""Generate a sample products XLS for testing the vendor bulk import feature.

Usage:
    python generate_products.py [--count 10] [--output products.xlsx]
"""

import argparse
import random
import string
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Columns match CreateProductRequest / XLS import format
HEADERS = ["name", "sku", "price", "description", "category", "imageUrl", "initialQuantity"]

# Category-specific product templates and real Unsplash photo IDs
CATEGORIES = {
    "Electronics": {
        "products": [
            ("Wireless Noise-Cancelling Headphones", "Premium over-ear headphones with 30hr battery life"),
            ("Mechanical Gaming Keyboard", "RGB backlit, tactile switches, full-size layout"),
            ("4K Webcam", "Ultra-HD webcam with built-in microphone and autofocus"),
            ("Portable SSD 1TB", "USB-C external solid-state drive, 1050MB/s read speed"),
            ("Smart LED Desk Lamp", "Adjustable color temperature with wireless charging pad"),
            ("USB-C Hub 7-in-1", "HDMI 4K, 3x USB-A, SD card reader, 100W PD"),
            ("Ergonomic Vertical Mouse", "Reduces wrist strain, 6 programmable buttons"),
            ("Laptop Stand Adjustable", "Aluminum foldable stand, fits 10-17 inch laptops"),
            ("Bluetooth Speaker Waterproof", "360° sound, IPX7 waterproof, 24hr battery"),
            ("Action Camera 4K60fps", "Waterproof action cam with image stabilization"),
        ],
        "photo_ids": [
            "1498049794561-7780e7231661",
            "1593642632559-0c6d3fc62b89",
            "1546054454-aa26e2b734c7",
            "1526570207772-4cd16ef5f8b5",
            "1588872657578-7efd1f1555ed",
        ],
        "price_range": (29.99, 299.99),
        "qty_range": (10, 200),
    },
    "Clothing": {
        "products": [
            ("Classic Crew-Neck T-Shirt", "100% organic cotton, pre-shrunk, unisex fit"),
            ("Slim-Fit Chino Pants", "Stretch cotton blend, available in 6 colours"),
            ("Waterproof Hiking Jacket", "3-layer shell, sealed seams, pit-zip vents"),
            ("Running Sneakers", "Lightweight mesh upper, responsive foam midsole"),
            ("Merino Wool Hoodie", "Ultra-soft, temperature-regulating, no itch"),
            ("High-Waist Yoga Leggings", "4-way stretch, moisture-wicking, squat-proof"),
            ("Oxford Button-Down Shirt", "Wrinkle-resistant cotton, regular fit"),
            ("Winter Puffer Coat", "Water-resistant shell, 600-fill down insulation"),
            ("Leather Bifold Wallet", "Genuine full-grain leather, RFID-blocking"),
            ("Canvas Tote Bag", "Heavyweight 12oz canvas, internal zip pocket"),
        ],
        "photo_ids": [
            "1523381210434-271e8be1f52b",
            "1542291026-7eec264c27ff",
            "1578681994506-b8f463449011",
            "1562572159-4efd6f5f5e8d",
            "1551232864-3f0890e1777c",
        ],
        "price_range": (19.99, 149.99),
        "qty_range": (20, 500),
    },
    "Home & Kitchen": {
        "products": [
            ("Pour-Over Coffee Maker", "Borosilicate glass carafe, BPA-free, 600ml"),
            ("High-Speed Blender 1800W", "Variable speed, tamper, 2L pitcher, self-clean"),
            ("Ceramic Non-Stick Pan 28cm", "PTFE-free coating, oven-safe to 260°C"),
            ("Stainless Steel Kettle 1.7L", "Temperature control, keep-warm, 360° base"),
            ("Air Fryer 5.5L Digital", "8 presets, dual-layer rack, dishwasher-safe basket"),
            ("Bamboo Cutting Board Set", "3-piece, juice grooves, anti-slip feet"),
            ("Cast Iron Dutch Oven 5.5qt", "Enamelled interior, suitable for induction"),
            ("Espresso Machine Semi-Auto", "15-bar pump, steam wand, 1.5L water tank"),
            ("Kitchen Scale Digital", "0.1g precision, tare function, stainless platform"),
            ("Vacuum Food Sealer", "Dry/moist modes, includes 10 bags"),
        ],
        "photo_ids": [
            "1556909114-f6e7ad7d3136",
            "1484101403633-562f891dc89a",
            "1556909045-b2b7e87b9ef8",
            "1542853878-e12dfdc17890",
            "1615484477778-ca3b77940c25",
        ],
        "price_range": (24.99, 249.99),
        "qty_range": (5, 150),
    },
    "Books": {
        "products": [
            ("Clean Code", "A handbook of agile software craftsmanship by Robert C. Martin"),
            ("Designing Data-Intensive Applications", "The big ideas behind reliable, scalable, maintainable systems"),
            ("The Pragmatic Programmer", "Your journey to mastery, 20th anniversary edition"),
            ("Atomic Habits", "Tiny changes, remarkable results — James Clear"),
            ("Deep Work", "Rules for focused success in a distracted world — Cal Newport"),
            ("System Design Interview Vol. 2", "In-depth case studies for senior engineers"),
            ("Domain-Driven Design", "Tackling complexity in the heart of software — Evans"),
            ("The Art of War", "Sun Tzu's classic strategy, modern translation"),
            ("Sapiens", "A brief history of humankind — Yuval Noah Harari"),
            ("Zero to One", "Notes on startups, or how to build the future — Peter Thiel"),
        ],
        "photo_ids": [
            "1524995997946-a1c2e315a42f",
            "1507842217343-583bb7270b66",
            "1532012197267-da84d127e765",
            "1544947950-fa07a98d237f",
        ],
        "price_range": (12.99, 49.99),
        "qty_range": (30, 300),
    },
    "Sports": {
        "products": [
            ("Yoga Mat Extra-Thick 8mm", "Non-slip natural rubber, alignment lines, carry strap"),
            ("Adjustable Dumbbell Set 2-24kg", "Quick-select dial, replaces 15 sets"),
            ("Resistance Bands Set 5-piece", "Heavy-duty latex, 10–50 lbs, includes door anchor"),
            ("Running Water Bottle 750ml", "BPA-free Tritan, leak-proof flip lid"),
            ("Foam Roller Deep Tissue", "High-density EVA, 33cm, trigger-point grid"),
            ("Jump Rope Speed Cable", "Ball-bearing handles, adjustable steel cable"),
            ("Gym Gloves with Wrist Wrap", "Anti-slip leather palm, ventilated, XS-XL"),
            ("Pull-Up Bar Doorframe", "No-screws, 300 lb capacity, multi-grip"),
            ("Cycling Helmet Urban", "MIPS rotational protection, 18 vents, CE certified"),
            ("Swimming Goggles Anti-Fog", "UV protection, wide-view, silicone seal"),
        ],
        "photo_ids": [
            "1571019614242-c5c5dee9f50b",
            "1535131749006-b7f58c99034b",
            "1517963879433-6ad2a04bb8a9",
            "1576678927484-cc907957088c",
        ],
        "price_range": (14.99, 199.99),
        "qty_range": (15, 400),
    },
    "Beauty": {
        "products": [
            ("Hyaluronic Acid Serum 30ml", "2% HA complex, vegan, fragrance-free"),
            ("Vitamin C Brightening Moisturiser", "SPF 30, 15% ascorbic acid, 50ml"),
            ("Retinol Night Cream 50ml", "0.3% retinol, encapsulated formula, reduces fine lines"),
            ("Micellar Cleansing Water 400ml", "No-rinse makeup remover, suitable for sensitive skin"),
            ("Niacinamide Toner 200ml", "10% niacinamide + zinc, minimises pores"),
            ("Natural Lip Balm Set 6-pack", "SPF 15, beeswax, various flavours"),
            ("Ionic Hair Dryer 1875W", "Tourmaline-infused, 3 heat / 2 speed, cool-shot"),
            ("Jade Facial Roller", "Genuine nephrite jade, dual-head, reduces puffiness"),
            ("Sunscreen SPF50+ 100ml", "Invisible finish, PA++++, reef-safe formula"),
            ("Charcoal Face Mask 100ml", "Deep-cleanse, kaolin clay + activated charcoal"),
        ],
        "photo_ids": [
            "1522335789203-aabd1fc54bc9",
            "1598440947619-2c35fc9aa908",
            "1596462502278-27bfdc403348",
            "1556228453-efd6c1ff04f6",
        ],
        "price_range": (9.99, 89.99),
        "qty_range": (25, 500),
    },
    "Garden": {
        "products": [
            ("Self-Watering Planter 30cm", "Sub-irrigation reservoir, terracotta-look HDPE"),
            ("Stainless Pruning Shears", "Bypass blades, ergonomic grip, safety lock"),
            ("Garden Kneeler & Seat", "Converts between kneeler and bench, tool pockets"),
            ("Expandable Garden Hose 50ft", "Latex inner, tangle-free, brass fittings"),
            ("Soil pH & Moisture Meter", "3-in-1 sensor, no batteries required"),
            ("Raised Bed Planter Kit 4x4ft", "Cedar wood, pre-cut, easy assembly"),
            ("Composting Bin 330L", "UV-resistant plastic, secure lid, air vents"),
            ("Garden Gloves Bamboo M/L", "Touchscreen-compatible fingertips, thorn-resistant"),
            ("Drip Irrigation Kit 50 plants", "Timer, adjustable flow emitters, UV-stable tubing"),
            ("Wildflower Seed Mix 100g", "25 native species, attracts pollinators"),
        ],
        "photo_ids": [
            "1416879595882-3373a0480b5b",
            "1585320806297-9794b3e4eeae",
            "1558618666-fcd25c85cd64",
            "1592861956120-e524fc739696",
        ],
        "price_range": (12.99, 129.99),
        "qty_range": (10, 200),
    },
    "Automotive": {
        "products": [
            ("Dash Cam 4K Front & Rear", "Sony STARVIS sensor, built-in GPS, loop recording"),
            ("Portable Jump Starter 2000A", "12V lithium, starts up to 8L petrol / 6L diesel"),
            ("Tyre Inflator Digital 150PSI", "Cordless, auto shut-off, includes needle & nozzle"),
            ("Car Vacuum Cleaner 120W", "Wet & dry, HEPA filter, 5m cable"),
            ("Seat-Back Organiser", "Waterproof Oxford, 4 pockets, fits all headrests"),
            ("Wireless CarPlay Adapter", "Converts wired CarPlay to wireless, plug & play"),
            ("Snow Brush & Ice Scraper", "Extendable to 1.3m, foam handle, brass blade"),
            ("Car Phone Mount Magsafe", "15W wireless charge, strong suction, 360° rotate"),
            ("OBD2 Bluetooth Scanner", "Reads & clears fault codes, live data via app"),
            ("All-Weather Floor Mats Set", "Custom laser-cut TPE, odourless, 4-piece"),
        ],
        "photo_ids": [
            "1492144534655-ae79c964c9d7",
            "1486496146582-9ffcd0b2b2b7",
            "1568605117036-5fe5e7bab0b7",
            "1503376780353-7e6692767b70",
        ],
        "price_range": (19.99, 199.99),
        "qty_range": (5, 100),
    },
}


def _sku(category: str, index: int) -> str:
    prefix = category[:3].upper()
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"{prefix}-{index:04d}-{suffix}"


def _image_url(photo_id: str) -> str:
    return f"https://images.unsplash.com/photo-{photo_id}?w=400&h=400&fit=crop"


def generate(count: int, output: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = [40, 20, 12, 60, 20, 70, 20]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    categories = list(CATEGORIES.keys())
    used_skus: set[str] = set()
    row_num = 2

    for i in range(count):
        cat_name = categories[i % len(categories)]
        cat = CATEGORIES[cat_name]
        prod_list = cat["products"]
        product = prod_list[i % len(prod_list)]
        name, description = product

        # Guarantee unique SKU
        while True:
            sku = _sku(cat_name, row_num)
            if sku not in used_skus:
                used_skus.add(sku)
                break

        price = round(random.uniform(*cat["price_range"]), 2)
        qty = random.randint(*cat["qty_range"])
        photo_id = random.choice(cat["photo_ids"])
        image_url = _image_url(photo_id)

        ws.append([name, sku, price, description, cat_name, image_url, qty])
        row_num += 1

    wb.save(output)
    print(f"Generated {count} product(s) → {Path(output).resolve()}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate sample products XLS for OMS vendor import")
    parser.add_argument("--count", type=int, default=10, help="Number of product rows to generate (default: 10)")
    parser.add_argument("--output", default="products.xlsx", help="Output file path (default: products.xlsx)")
    args = parser.parse_args()

    if args.count < 1:
        parser.error("--count must be at least 1")

    generate(args.count, args.output)


if __name__ == "__main__":
    main()
